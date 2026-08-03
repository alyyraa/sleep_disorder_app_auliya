"""Shared openpyxl report workbook generator."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils.timezone import format_indonesian_date, jakarta_now

BLUE = "0B3A75"


def build_excel(report_data, output):
    generated_at = report_data.get("generated_at") or jakarta_now()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = report_data["title"][:31]
    column_count = len(report_data["headers"])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    title_cell = sheet.cell(1, 1, report_data["title"])
    title_cell.font = Font(bold=True, size=14, color=BLUE)
    title_cell.alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=column_count)
    sheet.cell(2, 1, f"Tanggal Cetak: {format_indonesian_date(generated_at, include_time=True)} | Total Data: {report_data['total_records']}").alignment = Alignment(horizontal="center")
    row_index = 4
    if report_data["summary"]:
        for label, value in report_data["summary"]:
            sheet.cell(row_index, 1, label).font = Font(bold=True)
            sheet.cell(row_index, 2, value)
            row_index += 1
        row_index += 1
    header_row = row_index
    fill = PatternFill("solid", fgColor=BLUE)
    thin = Side(style="thin", color="A8BDD4")
    for column, header in enumerate(report_data["headers"], start=1):
        cell = sheet.cell(header_row, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in report_data["rows"]:
        row_index += 1
        for column, value in enumerate(row, start=1):
            cell = sheet.cell(row_index, column, value)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = f"A{header_row + 1}"
    for column in range(1, column_count + 1):
        width = max(len(str(sheet.cell(row, column).value or "")) for row in range(1, sheet.max_row)) + 2
        sheet.column_dimensions[get_column_letter(column)].width = min(max(width, 12), 35)
    workbook.save(output)
